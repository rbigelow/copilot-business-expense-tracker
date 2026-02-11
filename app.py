from flask import Flask, render_template, request, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime, timedelta
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import base64

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///expenses.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)

# Models
class Expense(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    date = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    category = db.Column(db.String(100), nullable=False)
    description = db.Column(db.String(200), nullable=False)
    amount = db.Column(db.Float, nullable=False)
    
    def to_dict(self):
        return {
            'id': self.id,
            'date': self.date.strftime('%Y-%m-%d'),
            'category': self.category,
            'description': self.description,
            'amount': self.amount
        }

# Routes
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/expenses', methods=['GET', 'POST'])
def expenses():
    if request.method == 'POST':
        data = request.json
        expense = Expense(
            date=datetime.strptime(data['date'], '%Y-%m-%d'),
            category=data['category'],
            description=data['description'],
            amount=float(data['amount'])
        )
        db.session.add(expense)
        db.session.commit()
        return jsonify(expense.to_dict()), 201
    else:
        expenses = Expense.query.order_by(Expense.date.desc()).all()
        return jsonify([e.to_dict() for e in expenses])

@app.route('/api/expenses/<int:id>', methods=['PUT', 'DELETE'])
def expense_detail(id):
    expense = Expense.query.get_or_404(id)
    
    if request.method == 'PUT':
        data = request.json
        expense.date = datetime.strptime(data['date'], '%Y-%m-%d')
        expense.category = data['category']
        expense.description = data['description']
        expense.amount = float(data['amount'])
        db.session.commit()
        return jsonify(expense.to_dict())
    
    elif request.method == 'DELETE':
        db.session.delete(expense)
        db.session.commit()
        return '', 204

@app.route('/api/report-data/<period>')
def report_data(period):
    """Get aggregated expense data for charts"""
    now = datetime.utcnow()
    
    # Calculate date range based on period
    if period == '30days':
        start_date = now - timedelta(days=30)
    elif period == '3months':
        start_date = now - timedelta(days=90)
    elif period == '6months':
        start_date = now - timedelta(days=180)
    elif period == '1year':
        start_date = now - timedelta(days=365)
    else:
        return jsonify({'error': 'Invalid period'}), 400
    
    # Get expenses for the period
    expenses = Expense.query.filter(Expense.date >= start_date).all()
    
    # Aggregate by category
    category_totals = {}
    for expense in expenses:
        if expense.category in category_totals:
            category_totals[expense.category] += expense.amount
        else:
            category_totals[expense.category] = expense.amount
    
    # Aggregate by date (for line chart)
    date_totals = {}
    for expense in expenses:
        date_key = expense.date.strftime('%Y-%m-%d')
        if date_key in date_totals:
            date_totals[date_key] += expense.amount
        else:
            date_totals[date_key] = expense.amount
    
    # Sort dates
    sorted_dates = sorted(date_totals.items())
    
    return jsonify({
        'categoryData': {
            'labels': list(category_totals.keys()),
            'values': list(category_totals.values())
        },
        'dateData': {
            'labels': [d[0] for d in sorted_dates],
            'values': [d[1] for d in sorted_dates]
        },
        'total': sum(category_totals.values()),
        'count': len(expenses)
    })

@app.route('/export/excel/<period>')
def export_excel(period):
    """Export expenses to Excel format"""
    now = datetime.utcnow()
    
    # Calculate date range
    if period == '30days':
        start_date = now - timedelta(days=30)
        period_name = 'Last 30 Days'
    elif period == '3months':
        start_date = now - timedelta(days=90)
        period_name = 'Last 3 Months'
    elif period == '6months':
        start_date = now - timedelta(days=180)
        period_name = 'Last 6 Months'
    elif period == '1year':
        start_date = now - timedelta(days=365)
        period_name = 'Last 1 Year'
    else:
        return jsonify({'error': 'Invalid period'}), 400
    
    # Get expenses
    expenses = Expense.query.filter(Expense.date >= start_date).order_by(Expense.date).all()
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Expenses'
    
    # Add title
    ws['A1'] = f'Business Expenses Report - {period_name}'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:E1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Add headers
    headers = ['Date', 'Category', 'Description', 'Amount']
    ws.append([])
    ws.append(headers)
    for cell in ws[3]:
        cell.font = Font(bold=True)
    
    # Add data
    total = 0
    for expense in expenses:
        ws.append([
            expense.date.strftime('%Y-%m-%d'),
            expense.category,
            expense.description,
            expense.amount
        ])
        total += expense.amount
    
    # Add total
    ws.append([])
    ws.append(['', '', 'Total:', total])
    last_row = ws.max_row
    ws[f'C{last_row}'].font = Font(bold=True)
    ws[f'D{last_row}'].font = Font(bold=True)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 12
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'expenses_{period}.xlsx'
    )

@app.route('/export/pdf/<period>')
def export_pdf(period):
    """Export expenses to PDF format with charts"""
    now = datetime.utcnow()
    
    # Calculate date range
    if period == '30days':
        start_date = now - timedelta(days=30)
        period_name = 'Last 30 Days'
    elif period == '3months':
        start_date = now - timedelta(days=90)
        period_name = 'Last 3 Months'
    elif period == '6months':
        start_date = now - timedelta(days=180)
        period_name = 'Last 6 Months'
    elif period == '1year':
        start_date = now - timedelta(days=365)
        period_name = 'Last 1 Year'
    else:
        return jsonify({'error': 'Invalid period'}), 400
    
    # Get expenses
    expenses = Expense.query.filter(Expense.date >= start_date).order_by(Expense.date).all()
    
    # Create PDF
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=letter)
    story = []
    styles = getSampleStyleSheet()
    
    # Title
    title = Paragraph(f'<b>Business Expenses Report - {period_name}</b>', styles['Title'])
    story.append(title)
    story.append(Spacer(1, 0.3*inch))
    
    # Summary statistics
    total_amount = sum(e.amount for e in expenses)
    summary = Paragraph(f'Total Expenses: ${total_amount:.2f} | Number of Transactions: {len(expenses)}', styles['Normal'])
    story.append(summary)
    story.append(Spacer(1, 0.3*inch))
    
    # Create category chart
    if expenses:
        category_totals = {}
        for expense in expenses:
            category_totals[expense.category] = category_totals.get(expense.category, 0) + expense.amount
        
        # Create pie chart
        fig, ax = plt.subplots(figsize=(6, 4))
        ax.pie(category_totals.values(), labels=category_totals.keys(), autopct='%1.1f%%', startangle=90)
        ax.set_title('Expenses by Category')
        
        # Save chart to BytesIO
        chart_buffer = BytesIO()
        plt.savefig(chart_buffer, format='png', bbox_inches='tight')
        chart_buffer.seek(0)
        plt.close()
        
        # Add chart to PDF
        chart_img = Image(chart_buffer, width=4*inch, height=3*inch)
        story.append(chart_img)
        story.append(Spacer(1, 0.3*inch))
    
    # Expense table
    table_data = [['Date', 'Category', 'Description', 'Amount']]
    for expense in expenses:
        table_data.append([
            expense.date.strftime('%Y-%m-%d'),
            expense.category,
            expense.description,
            f'${expense.amount:.2f}'
        ])
    
    # Add total row
    table_data.append(['', '', 'Total:', f'${total_amount:.2f}'])
    
    # Create table
    table = Table(table_data, colWidths=[1.2*inch, 1.5*inch, 3*inch, 1*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (3, 0), (3, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(table)
    
    # Build PDF
    doc.build(story)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'expenses_{period}.pdf'
    )

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(debug=True)
